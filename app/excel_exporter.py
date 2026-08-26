"""Excel 导出 — 将 JD + AI 分析结果保存为 Excel 文件

导出内容：
- 时间戳
- 岗位名称、公司、薪资、城市
- 岗位描述(JD)、任职要求
- AI 匹配分数、是否匹配、匹配理由
- AI 优势、劣势、建议打招呼消息
- AI 完整返回 JSON（用于评测）
- AI 响应耗时（秒）
- 使用的提示词版本号
- 投递状态（已投递/已跳过）
"""
import os
import time
import threading
import json
from typing import Optional

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

DATA_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "data")
JD_DIR = os.path.join(DATA_DIR, "jd_analysis")

# 表头样式
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="5b7cfa", end_color="5b7cfa", fill_type="solid")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)

# 匹配/不匹配行填充
MATCH_FILL = PatternFill(start_color="e8f5e9", end_color="e8f5e9", fill_type="solid")
SKIP_FILL = PatternFill(start_color="fff3e0", end_color="fff3e0", fill_type="solid")

_lock = threading.Lock()

# Excel 单元格最大字符限制
MAX_CELL_LENGTH = 32000


def _ensure_jd_dir():
    os.makedirs(JD_DIR, exist_ok=True)


def _get_workbook(path: str) -> openpyxl.Workbook:
    """加载现有工作簿或创建新的。"""
    if os.path.exists(path):
        return openpyxl.load_workbook(path)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "JD分析记录"
    _write_header(ws)
    return wb


def _write_header(ws):
    """写入表头。"""
    headers = [
        "时间", "岗位名称", "公司", "薪资", "城市",
        "岗位描述(JD)", "任职要求",
        "AI匹配分", "是否匹配", "匹配理由",
        "AI优势", "AI劣势", "建议打招呼",
        "AI响应耗时(s)", "提示词版本",
        "AI完整返回JSON", "投递状态", "岗位URL",
    ]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER


def _auto_width(ws):
    """自动调整列宽。"""
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                val = str(cell.value) if cell.value else ""
                # 中文字符算2个宽度
                width = sum(2 if ord(c) > 127 else 1 for c in val)
                if width > max_len:
                    max_len = width
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 4, 50)


def _truncate(text: str, max_len: int = MAX_CELL_LENGTH) -> str:
    """截断超长文本，避免 Excel 单元格溢出。"""
    if not text:
        return ""
    if len(text) > max_len:
        return text[:max_len] + "...[截断]"
    return text


def save_jd_analysis(
    job: dict,
    ai_result: Optional[dict] = None,
    skipped: bool = False,
    query: str = "",
    city: str = "",
    prompt_version: str = "",
    jd_description: str = "",
    jd_requirements: str = "",
    ai_duration: float = 0.0,
) -> str:
    """保存一条 JD + AI 分析记录到 Excel。

    Args:
        job: 岗位信息字典
        ai_result: AI 分析结果（可选）
        skipped: 是否跳过
        query: 搜索关键词
        city: 城市
        prompt_version: 提示词版本号
        jd_description: 岗位描述原文
        jd_requirements: 任职要求原文
        ai_duration: AI 响应耗时（秒）

    Returns:
        保存的文件路径

    Raises:
        PermissionError: 文件被占用时抛出
    """
    _ensure_jd_dir()
    with _lock:
        # 所有数据写入同一个文件
        filename = "JD分析记录.xlsx"
        filepath = os.path.join(JD_DIR, filename)

        wb = _get_workbook(filepath)
        ws = wb["JD分析记录"]

        # 追加行
        row_idx = ws.max_row + 1
        score = ai_result.get("score", "") if ai_result else ""
        is_match = "是" if (ai_result and ai_result.get("is_match")) else "否" if ai_result else ""
        reason = ai_result.get("reason", "") if ai_result else ""
        strengths = ", ".join(ai_result.get("strengths", [])) if ai_result else ""
        weaknesses = ", ".join(ai_result.get("weaknesses", [])) if ai_result else ""
        suggested = ai_result.get("suggested_greeting", "") if ai_result else ""
        status = "已跳过" if skipped else "已投递"

        # AI 完整返回 JSON（用于评测）
        ai_raw_json = json.dumps(ai_result, ensure_ascii=False) if ai_result else ""

        row_data = [
            time.strftime("%Y-%m-%d %H:%M:%S"),
            job.get("job_name", ""),
            job.get("company", ""),
            job.get("salary", ""),
            city,
            _truncate(jd_description),
            _truncate(jd_requirements),
            score,
            is_match,
            reason,
            strengths,
            weaknesses,
            suggested,
            round(ai_duration, 2) if ai_duration else "",
            prompt_version,
            _truncate(ai_raw_json, 10000),  # JSON 限制 10000 字符
            status,
            job.get("url", ""),
        ]

        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(wrap_text=True, vertical="center")
            if status == "已投递":
                cell.fill = MATCH_FILL
            elif status == "已跳过":
                cell.fill = SKIP_FILL

        _auto_width(ws)
        try:
            wb.save(filepath)
        except PermissionError:
            # 文件被占用，尝试用新文件名
            ts = time.strftime("%Y%m%d_%H%M%S")
            alt_filename = f"JD分析记录_{ts}.xlsx"
            alt_filepath = os.path.join(JD_DIR, alt_filename)
            wb.save(alt_filepath)
            return alt_filepath
        return filepath


def get_analysis_files() -> list:
    """获取所有分析 Excel 文件列表。"""
    _ensure_jd_dir()
    files = []
    for f in os.listdir(JD_DIR):
        if f.endswith(".xlsx"):
            path = os.path.join(JD_DIR, f)
            files.append({
                "name": f,
                "path": path,
                "size": os.path.getsize(path),
                "mtime": os.path.getmtime(path),
            })
    files.sort(key=lambda x: x["mtime"], reverse=True)
    return files


def export_all_from_chats_log() -> int:
    """从 chats_log.json 导出所有历史记录到 Excel。

    Returns:
        导出的记录数
    """
    chats_log_file = os.path.join(DATA_DIR, "chats_log.json")
    if not os.path.exists(chats_log_file):
        return 0

    _ensure_jd_dir()
    with _lock:
        filename = "JD分析记录.xlsx"
        filepath = os.path.join(JD_DIR, filename)

        # 如果文件已存在，先读取已有的时间戳+岗位名做去重
        existing_keys = set()
        if os.path.exists(path=filepath):
            try:
                wb_old = openpyxl.load_workbook(filepath, read_only=True)
                ws_old = wb_old.active
                for row in ws_old.iter_rows(min_row=2, values_only=True):
                    if row[0] and row[1]:
                        existing_keys.add(f"{row[0]}_{row[1]}")
                wb_old.close()
            except Exception:
                pass

        wb = _get_workbook(filepath)
        ws = wb["JD分析记录"]

        with open(chats_log_file, "r", encoding="utf-8") as f:
            logs = json.load(f)

        # 过滤掉非 dict 的旧格式记录
        logs = [log for log in logs if isinstance(log, dict)]

        count = 0
        for log in logs:
            key = f"{log.get('time', '')}_{log.get('job_name', '')}"
            if key in existing_keys:
                continue
            existing_keys.add(key)

            row_idx = ws.max_row + 1
            ai_result = log.get("ai_result")
            score = ai_result.get("score", "") if ai_result else ""
            is_match = "是" if (ai_result and ai_result.get("is_match")) else "否" if ai_result else ""
            reason = ai_result.get("reason", "") if ai_result else ""
            strengths = ", ".join(ai_result.get("strengths", [])) if ai_result else ""
            weaknesses = ", ".join(ai_result.get("weaknesses", [])) if ai_result else ""
            suggested = ai_result.get("suggested_greeting", "") if ai_result else ""
            status = "已跳过" if log.get("skipped") else "已投递"
            ai_raw_json = json.dumps(ai_result, ensure_ascii=False) if ai_result else ""

            row_data = [
                log.get("time", ""),
                log.get("job_name", ""),
                log.get("company", ""),
                log.get("salary", ""),
                log.get("city", ""),
                _truncate(log.get("jd_description", "")),
                _truncate(log.get("jd_requirements", "")),
                score,
                is_match,
                reason,
                strengths,
                weaknesses,
                suggested,
                round(log.get("ai_duration", 0), 2),
                log.get("prompt_version", ""),
                _truncate(ai_raw_json, 10000),
                status,
                log.get("url", ""),
            ]

            for col_idx, val in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.border = THIN_BORDER
                cell.alignment = Alignment(wrap_text=True, vertical="center")
                if status == "已投递":
                    cell.fill = MATCH_FILL
                elif status == "已跳过":
                    cell.fill = SKIP_FILL

            count += 1

        if count > 0:
            _auto_width(ws)
            try:
                wb.save(filepath)
            except PermissionError:
                ts = time.strftime("%Y%m%d_%H%M%S")
                alt_filepath = os.path.join(JD_DIR, f"JD分析记录_{ts}.xlsx")
                wb.save(alt_filepath)

        return count
